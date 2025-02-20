from flask import Flask, request, jsonify, render_template, redirect
import datetime
import smtplib
from email.mime.text import MIMEText
import openpyxl
import requests
import os

app = Flask(__name__)

# GitHub raw file link (replace with actual URL)
GITHUB_FILE_URL = "https://github.com/aryankhetarpal/Machine-Breakdown/blob/main/machine_breakdowns.xlsx"

def fetch_and_open_excel():
    try:
        # Step 1: Download the Excel file
        response = requests.get(GITHUB_FILE_URL, stream=True)

        # Step 2: Check if download is successful
        if response.status_code != 200:
            return f"Error: Failed to fetch the file, Status Code: {response.status_code}"

        # Step 3: Check if response content type is an Excel file
        content_type = response.headers.get("Content-Type", "")
        if "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" not in content_type:
            return f"Error: Expected an Excel file but received {content_type}"

        # Step 4: Save and open the file
        temp_file = "temp_machine_breakdowns.xlsx"
        with open(temp_file, "wb") as file:
            file.write(response.content)

        # Step 5: Open the file with openpyxl
        wb = openpyxl.load_workbook(temp_file)
        ws = wb.active
        return f"Excel file opened successfully! First cell: {ws.cell(row=1, column=1).value}"

    except Exception as e:
        return f"Error: {str(e)}"

# Email configuration (Change these to your details)
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587
EMAIL_SENDER = "your-email@gmail.com"
EMAIL_PASSWORD = "your-email-password"
EMAIL_MANAGER = "laxmi@pck-buderus.com"

# Function to send an email alert
def send_email_alert(machine_name):
    subject = f"ALERT: Frequent Breakdown for {machine_name}"
    body = f"The machine '{machine_name}' has encountered more than 2 breakdowns within the last week."
    
    msg = MIMEText(body)
    msg["Subject"] = subject
    msg["From"] = EMAIL_SENDER
    msg["To"] = EMAIL_MANAGER

    try:
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(EMAIL_SENDER, EMAIL_PASSWORD)
            server.sendmail(EMAIL_SENDER, EMAIL_MANAGER, msg.as_string())
        print("Email sent successfully.")
    except Exception as e:
        print("Error sending email:", e)

# Route for main form page
@app.route("/")
def index():
    return render_template("index.html")

# Handle form submission
@app.route("/submit", methods=["POST"])
def submit():
    try:
        machine_name = request.form.get("machine_name")
        issue = request.form.get("issue")
        date = datetime.datetime.now().strftime("%Y-%m-%d")

        if not machine_name or not issue:
            return jsonify({"error": "Both fields are required"}), 400

        # Download the Excel file
        response = requests.get(GITHUB_FILE_URL)
        if response.status_code != 200:
            return jsonify({"error": "Failed to fetch the Excel file"}), 500

        # Save file temporarily and open it
        temp_file = "temp_machine_breakdowns.xlsx"
        with open(temp_file, "wb") as file:
            file.write(response.content)

        wb = openpyxl.load_workbook(temp_file)
        ws = wb.active
        ws.append([date, machine_name, issue])
        wb.save(temp_file)

        # Check if the machine has more than 2 issues in a week
        recent_issues = 0
        one_week_ago = datetime.datetime.now() - datetime.timedelta(days=7)

        for row in ws.iter_rows(values_only=True):
            if row[0] and row[1]:  # Ensure row has valid data
                try:
                    row_date = datetime.datetime.strptime(row[0], "%Y-%m-%d")
                    if row[1] == machine_name and row_date >= one_week_ago:
                        recent_issues += 1
                except ValueError:
                    continue  # Skip invalid date formats

        if recent_issues > 2:
            send_email_alert(machine_name)

        return jsonify({"message": "Breakdown recorded successfully!"})

    except Exception as e:
        return jsonify({"error": str(e)}), 500

# Route to open Excel file (redirects to GitHub)
@app.route("/open_log")
def open_log():
    return redirect(GITHUB_FILE_URL)

if __name__ == "__main__":
    app.run(debug=True)
